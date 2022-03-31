
from flask import Flask, render_template, request,jsonify
import requests
from bs4 import BeautifulSoup
import pandas as pdd
import numpy as np
import re
import string
import nltk
from nltk.corpus import stopwords
from nltk import PorterStemmer
import matplotlib.pyplot as plt
import seaborn as sns
from wordcloud import WordCloud
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from collections import Counter
from matplotlib import rcParams
import os

import pickle
import flasgger
from flasgger import Swagger
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.naive_bayes import MultinomialNB
import joblib

import uuid
def my_random_string(string_length=6):
    random = str(uuid.uuid4())
    random = random.replace("-","")
    return random[0:string_length]

app = Flask(__name__)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 300
mnb = pickle.load(open('Naive_Bayes_model.pkl','rb'))
countVect = pickle.load(open('countVect.pkl','rb'))

#Method Arena 
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pdd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
reviewlist = []

    
def get_soup(url):
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')
    return soup


def get_reviews(soup):
    reviews = soup.find_all('div', {'data-hook': 'review'})
    try:
        for item in reviews:
            review = {
            'Product': soup.title.text.replace('Amazon.in:Customer reviews:', '').strip(),
            'Customer Name':item.find('span',class_='a-profile-name').text.strip(),
            'Review Title': item.find('a', {'data-hook': 'review-title'}).text.strip(),
            'Rating':  float(item.find('i', {'data-hook': 'review-star-rating'}).text.replace('out of 5 stars', '').strip()),
            'Reviews': item.find('span', {'data-hook': 'review-body'}).text.strip(),
            }
            reviewlist.append(review)
    except:
        pass


#End Method Arena 
        
@app.after_request
def add_header(response):
# response.cache_control.no_store = True
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

@app.route("/") 
def index(): 
    return(render_template("index.html"))

processed_text=""
STOPWORDS=""
@app.route('/', methods=['POST'])
def my_form_post():
    import pandas as pd
    processed_text = request.form['text'] 
    print(processed_text)
    processed_text=processed_text.replace("dp","product-reviews",1)
    fi = my_random_string(6)
    print("Temporary Variable : " + fi)

    for x in range(1,30):
        pg=str(x)
        url_main=processed_text
        url_split1="/ref=cm_cr_arp_d_paging_btm_next_"
        url_split2="?ie=UTF8&reviewerType=all_reviews&pageNumber="
        
        url_final = url_main+url_split1+pg+url_split2+pg
        print(url_final) 
        soup = get_soup(url_final)
        print(f'Getting page: {x}')
        get_reviews(soup)
        print(len(reviewlist))
        if not soup.find('li', {'class': 'a-disabled a-last'}):
            pass
        else:
            break
    
    df = pd.DataFrame(reviewlist)
    append_df_to_excel('static/'+fi+'.xlsx', df, header=True, index=False)
    print('End.')
    
    df=pd.read_excel(r'static/'+fi+'.xlsx')
    
    
    x=[]
    x=df.Reviews
    
    REPLACE_NO_SPACE = re.compile("[.;:!\'?,\"()\[\]]")
    REPLACE_WITH_SPACE = re.compile("(<br\s*/><br\s*/>)|(\-)|(\/)")
    
    def preprocess_reviews(reviews):
        reviews = [REPLACE_NO_SPACE.sub("", line.lower()) for line in reviews]
        reviews = [REPLACE_WITH_SPACE.sub(" ", line) for line in reviews]
        
        return reviews
    print('153')

    reviews_train_clean = preprocess_reviews(x)
    df['Cleaned Reviews']=reviews_train_clean
    
    STOPWORDS=stopwords.words("english")
    def deEmojify(inputString):
        return inputString.encode('ascii', 'ignore').decode('ascii')
        
    wordcloud = WordCloud(height=500, width=500)
    wordcloud = wordcloud.generate(' '.join(df['Cleaned Reviews'].tolist()))
    plt.imshow(wordcloud)
    plt.title("Most common words in the reviews")
    plt.axis('off')
    plt.show()
    wordcloud =  wordcloud.to_file('static/wordcloud.png')
    analyser = SentimentIntensityAnalyzer()
    print('170')
    def sentiment_analyzer_scores(sentence):
        score = analyser.polarity_scores(sentence)
        return score
        
    def compound_score(text):
        comp=sentiment_analyzer_scores(text)
        return comp['compound']
        
    df['sentiment_score']=df['Cleaned Reviews'].apply(lambda x:compound_score(x))
    def sentiment_category(score):
        if score >= 0.05:
            return "positive"
        elif score <= -0.05:
            return "negative"
        else:
            return "neutral"
            
    df['review_category']=df['sentiment_score'].apply(lambda x:sentiment_category(x))
    sns.countplot(df['review_category']).set_title("Distribution of Reviews Category")
    plt.savefig('static/count_plot.png')
    
    positive_reviews=df.loc[df['review_category']=='positive','Cleaned Reviews'].tolist()
    negative_reviews=df.loc[df['review_category']=='negative','Cleaned Reviews'].tolist()
    # print('194')
    #
    #POSITIVE AND NEGATIVE WORD CLOUD ARENA
    #

    def getMostCommon(reviews_list,topn=20):
        reviews=" ".join(reviews_list)
        tokenised_reviews=reviews.split(" ")
        
        
        freq_counter=Counter(tokenised_reviews)
        return freq_counter.most_common(topn)
    
    def plotMostCommonWords(reviews_list,topn=20,title="Common Review Words",color="blue",axis=None): #default number of words is given as 20
        top_words=getMostCommon(reviews_list,topn=topn)
        data=pd.DataFrame()
        data['words']=[val[0] for val in top_words]
        data['freq']=[val[1] for val in top_words]
        if axis!=None:
            sns.barplot(y='words',x='freq',data=data,color=color,ax=axis).set_title(title+" top "+str(topn))
        else:
            sns.barplot(y='words',x='freq',data=data,color=color).set_title(title+" top "+str(topn))
    
    def generateNGram(text,n):
        tokens=text.split(" ")
        ngrams = zip(*[tokens[i:] for i in range(n)])
        return ["_".join(ngram) for ngram in ngrams]
    
    rcParams['figure.figsize'] = 14,10 ## Sets the heigth and width of image
    fig,ax=plt.subplots(1,2)
    fig.subplots_adjust(wspace=1.0) #Adjusts the space between the two plots
    plotMostCommonWords(positive_reviews,60,"Positive Review Unigrams",axis=ax[0])
    plotMostCommonWords(negative_reviews,60,"Negative Review Unigrams",color="red",axis=ax[1])
    plt.savefig('static/unigram.png')
    
    return render_template('index.html')

@app.route('/predict',methods=['POST'])
def predict():

    if request.method == 'POST':
        Reviews = request.form['Reviews']
        data = [Reviews]
        vect = countVect.transform(data).toarray()
        my_prediction = mnb.predict(vect)
    return render_template('index.html',prediction = my_prediction)

@app.route("/geti")
def get_p():
    return "count_plot.png"

@app.route("/getu")
def get_u():
    return "unigram.png"

@app.route("/getw")
def get_w():
    return "wordcloud.png"

app.config["CACHE_TYPE"] = "null"
app.run(debug=False)